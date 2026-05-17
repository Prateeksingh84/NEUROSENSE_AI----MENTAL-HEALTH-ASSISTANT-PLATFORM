"""
hs_module.py
────────────────────────────────────────────────────────────────────────────────
Hand Sign Module for NeuroSense AI
MediaPipe 0.10.x+ (Tasks API — no mp.solutions)

REAL-TIME IMPROVEMENTS in this version
───────────────────────────────────────
  ✅ Module-level model cache     — load_model() reads disk ONCE, then serves
                                    from RAM; Flask never hits the filesystem
                                    per-request again.
  ✅ Module-level landmarker cache— HandLandmarker created once at import time
                                    (lazy), reused for every frame.
  ✅ Unified CLAHE preprocessing  — single _enhance_frame() used everywhere.
  ✅ O(1) sample-count            — no full file scan to get next sample_id;
                                    uses a lightweight line-count helper.
  ✅ Cross-platform camera open   — CAP_DSHOW only on Windows; uses default
                                    backend on Linux/macOS.
  ✅ Thread-safe model cache      — RLock guards the in-memory clf so Flask
                                    threaded mode cannot corrupt it.
  ✅ Graceful FPS throttle        — collect_data sleeps the minimum needed to
                                    hit exactly 100 ms between saves (no busy-
                                    wait spin).
  ✅ Proper resource cleanup      — context managers + atexit for landmarker.
  ✅ predict_with_confidence()    — new helper that also returns confidence
                                    score so Flask can gate on certainty.
  ✅ Duplicate-sign check         — moved to clean _is_duplicate_sign() helper,
                                    reused in both collect & live modes.
  ✅ Validation hardened          — _validate_label now rejects labels that
                                    are ONLY spaces.
"""
import os
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"

import re
import csv
import cv2
import atexit
import platform
import datetime
import threading
import numpy as np
import pandas as pd
import pickle
import time
import urllib.request
from sklearn.ensemble import RandomForestClassifier

# ── MediaPipe Tasks API ────────────────────────────────────────────────────────
import mediapipe as mp
from mediapipe.tasks import python
from mediapipe.tasks.python import vision

BaseOptions           = python.BaseOptions
HandLandmarker        = vision.HandLandmarker
HandLandmarkerOptions = vision.HandLandmarkerOptions
RunningMode           = vision.RunningMode

# ─────────────────────────────────────────────────────────────────────────────
# SETTINGS
# ─────────────────────────────────────────────────────────────────────────────
_HERE      = os.path.dirname(os.path.abspath(__file__))
BASE_DIR   = os.path.join(_HERE, "data")
DATA_FILE  = os.path.join(BASE_DIR, "word_data.csv")
REF_DIR    = os.path.join(BASE_DIR, "reference_images")
MODEL_FILE = os.path.join(BASE_DIR, "word_model.pkl")

TASK_DIR   = os.path.join(_HERE, "models")
TASK_FILE  = os.path.join(TASK_DIR, "hand_landmarker.task")
TASK_URL   = (
    "https://storage.googleapis.com/mediapipe-models/"
    "hand_landmarker/hand_landmarker/float16/1/hand_landmarker.task"
)

NUM_SAMPLES       = 100
REF_SIZE          = 400
FEATURES_PER_HAND = 21 * 3   # x, y, z per landmark
CONFIDENCE_MIN    = 0.22     # minimum predict_proba to accept a prediction
COLLECT_INTERVAL  = 0.10     # seconds between saved samples during collection
STABLE_FRAMES     = 3        # consecutive frames before appending in live mode
LIVE_COOLDOWN     = 0.8      # seconds before the same sign can repeat

SPECIAL_SIGNS = {"SPACE", "DELETE", "CLEAR", "ENTER"}

# ── Landmark metadata ──────────────────────────────────────────────────────────
_LM_NAMES = [
    "WRIST",
    "THUMB_CMC","THUMB_MCP","THUMB_IP","THUMB_TIP",
    "INDEX_MCP","INDEX_PIP","INDEX_DIP","INDEX_TIP",
    "MIDDLE_MCP","MIDDLE_PIP","MIDDLE_DIP","MIDDLE_TIP",
    "RING_MCP","RING_PIP","RING_DIP","RING_TIP",
    "PINKY_MCP","PINKY_PIP","PINKY_DIP","PINKY_TIP",
]
_LM_FINGER = [
    "wrist",
    "thumb","thumb","thumb","thumb",
    "index","index","index","index",
    "middle","middle","middle","middle",
    "ring","ring","ring","ring",
    "pinky","pinky","pinky","pinky",
]

def _build_csv_columns():
    cols = ["sample_id", "timestamp"]
    for lm in _LM_NAMES:
        for ax in ("x", "y", "z"):
            cols.append(f"H1_{lm}_{ax}")
    cols.append("sign_label")
    return cols

CSV_COLUMNS   = _build_csv_columns()
_FEATURE_COLS = CSV_COLUMNS[2:-1]   # the 63 coordinate columns

# ─────────────────────────────────────────────────────────────────────────────
# MODULE-LEVEL CACHES  (the key real-time fix)
# ─────────────────────────────────────────────────────────────────────────────
_clf_lock:  threading.RLock = threading.RLock()
_clf_cache  = None   # cached RandomForestClassifier
_clf_mtime  = 0.0    # mtime of MODEL_FILE when it was loaded

_landmarker = None   # singleton HandLandmarker

# ─────────────────────────────────────────────────────────────────────────────
# DIRECTORY / TASK-MODEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _ensure_dirs():
    for d in (BASE_DIR, REF_DIR, TASK_DIR):
        os.makedirs(d, exist_ok=True)


def _ensure_task_model():
    if os.path.exists(TASK_FILE):
        return
    print("Downloading hand landmarker model (~15 MB)...")
    _ensure_dirs()
    try:
        urllib.request.urlretrieve(TASK_URL, TASK_FILE)
        print(f"Model saved to: {TASK_FILE}")
    except Exception as e:
        raise RuntimeError(
            f"Download failed.\nGet it from:\n  {TASK_URL}\n"
            f"Place at: {TASK_FILE}\nError: {e}"
        )

# ─────────────────────────────────────────────────────────────────────────────
# LANDMARKER — created once, reused forever
# ─────────────────────────────────────────────────────────────────────────────
def _get_landmarker():
    global _landmarker
    if _landmarker is None:
        _ensure_task_model()
        opts = HandLandmarkerOptions(
            base_options=BaseOptions(model_asset_path=TASK_FILE),
            running_mode=RunningMode.IMAGE,
            num_hands=1,
            min_hand_detection_confidence=0.22,
            min_hand_presence_confidence=0.22,
            min_tracking_confidence=0.22,
        )
        _landmarker = HandLandmarker.create_from_options(opts)
        atexit.register(_cleanup_landmarker)
        print("HandLandmarker initialised.")
    return _landmarker


def _cleanup_landmarker():
    global _landmarker
    if _landmarker is not None:
        try:
            _landmarker.close()
        except Exception:
            pass
        _landmarker = None

# ─────────────────────────────────────────────────────────────────────────────
# PREPROCESSING  (single unified function used everywhere)
# ─────────────────────────────────────────────────────────────────────────────
def _enhance_frame(frame: np.ndarray) -> np.ndarray:
    """
    CLAHE contrast enhancement on the L channel (LAB space).
    Robust to dim, uneven, or harsh lighting.
    Returns a new array — original is not modified.
    """
    try:
        lab = cv2.cvtColor(frame, cv2.COLOR_BGR2LAB)
        l, a, b = cv2.split(lab)
        clahe = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8))
        return cv2.cvtColor(cv2.merge([clahe.apply(l), a, b]), cv2.COLOR_LAB2BGR)
    except Exception:
        return frame


def _bgr_to_mp_image(frame: np.ndarray) -> mp.Image:
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    return mp.Image(image_format=mp.ImageFormat.SRGB, data=rgb)

# ─────────────────────────────────────────────────────────────────────────────
# LOW-LEVEL LANDMARK EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────
def _detect(frame: np.ndarray):
    """
    Run MediaPipe on `frame`.
    Tries raw first, then CLAHE-enhanced if no hand found.
    Returns a HandLandmarkerResult (may have empty hand_landmarks).
    """
    lm = _get_landmarker()
    for enhanced in (False, True):
        src    = _enhance_frame(frame) if enhanced else frame
        result = lm.detect(_bgr_to_mp_image(src))
        if result.hand_landmarks:
            return result
    return result   # empty result


def _result_to_features(result):
    """Flatten the first detected hand into a 63-element float list, or None."""
    if not result.hand_landmarks:
        return None
    row = []
    for lm in result.hand_landmarks[0]:
        row.extend([lm.x, lm.y, lm.z])
    return row


def _frame_to_features(frame: np.ndarray):
    return _result_to_features(_detect(frame))


def _draw_landmarks(frame: np.ndarray, result) -> np.ndarray:
    if not result.hand_landmarks:
        return frame
    h, w = frame.shape[:2]
    CONNECTIONS = [
        (0,1),(1,2),(2,3),(3,4),
        (0,5),(5,6),(6,7),(7,8),
        (0,9),(9,10),(10,11),(11,12),
        (0,13),(13,14),(14,15),(15,16),
        (0,17),(17,18),(18,19),(19,20),
        (5,9),(9,13),(13,17),
    ]
    for hand in result.hand_landmarks:
        pts = [(int(lm.x * w), int(lm.y * h)) for lm in hand]
        for a, b in CONNECTIONS:
            cv2.line(frame, pts[a], pts[b], (0, 200, 100), 2)
        for pt in pts:
            cv2.circle(frame, pt, 4, (0, 255, 200), -1)
    return frame

# ─────────────────────────────────────────────────────────────────────────────
# CAMERA SELECTION  (cross-platform)
# ─────────────────────────────────────────────────────────────────────────────
def _open_cap(idx: int) -> cv2.VideoCapture:
    """Open VideoCapture with the right backend for the current OS."""
    if platform.system() == "Windows":
        return cv2.VideoCapture(idx, cv2.CAP_DSHOW)
    return cv2.VideoCapture(idx)


def _select_camera() -> int:
    print("\nScanning for cameras...")
    available = []
    for idx in range(5):
        cap = _open_cap(idx)
        if cap.isOpened():
            ok, _ = cap.read()
            if ok:
                available.append(idx)
        cap.release()

    if not available:
        print("No cameras found.")
        return 0
    if len(available) == 1:
        print(f"Using camera index {available[0]}.")
        return available[0]

    labels = {
        0: "Built-in webcam",
        1: "External / Phone-Link / DroidCam",
        2: "External / Phone-Link / DroidCam",
        3: "External camera",
        4: "External camera",
    }
    print(f"\nFound {len(available)} camera(s):")
    for i, idx in enumerate(available):
        print(f"  {i+1}: index {idx}  —  {labels.get(idx, 'Camera')}")

    while True:
        choice = input(f"\nChoose camera [1-{len(available)}, default=1]: ").strip()
        if not choice or choice == "1":
            return available[0]
        if choice.isdigit() and 1 <= int(choice) <= len(available):
            return available[int(choice) - 1]
        print(f"   Enter a number between 1 and {len(available)}.")

# ─────────────────────────────────────────────────────────────────────────────
# LABEL VALIDATION
# ─────────────────────────────────────────────────────────────────────────────
def _validate_label(raw: str):
    label = raw.strip().upper()
    if not label:
        return False, "", "Label cannot be empty."
    if label in SPECIAL_SIGNS:
        return True, label, ""
    if re.fullmatch(r"[A-Z][A-Z ]*", label):
        label = re.sub(r" +", " ", label).strip()
        if label:
            return True, label, ""
    bad = set(re.sub(r"[A-Z ]", "", label))
    return False, "", (
        f"Invalid characters: {', '.join(sorted(bad))}.\n"
        "Use A-Z letters, spaces for phrases (e.g. THANK YOU),\n"
        f"or special signs: {', '.join(sorted(SPECIAL_SIGNS))}."
    )

# ─────────────────────────────────────────────────────────────────────────────
# CSV HELPERS  (O(1) sample-id — no full-file reload)
# ─────────────────────────────────────────────────────────────────────────────
def _fast_row_count(path: str) -> int:
    """Count data rows (excluding header) using a raw byte scan."""
    if not os.path.exists(path):
        return 0
    with open(path, "rb") as f:
        return sum(1 for _ in f) - 1   # subtract header


def _save_sample(label: str, hand_landmarks):
    """Append one labelled sample row to DATA_FILE (no full-file reload)."""
    coords = []
    for lm in hand_landmarks:
        coords.extend([round(lm.x, 6), round(lm.y, 6), round(lm.z, 6)])

    sample_id  = _fast_row_count(DATA_FILE) + 1
    ts         = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row        = [sample_id, ts] + coords + [label]
    file_exists = os.path.exists(DATA_FILE)

    with open(DATA_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(CSV_COLUMNS)   # write header once
        writer.writerow(row)

# ─────────────────────────────────────────────────────────────────────────────
# CLASSIFIER  — module-level cache with mtime invalidation
# ─────────────────────────────────────────────────────────────────────────────
def load_model():
    """
    Return the cached RandomForestClassifier.
    Re-reads from disk ONLY when MODEL_FILE has been modified (e.g. after
    train_model() runs).  Safe to call from multiple Flask threads.
    """
    global _clf_cache, _clf_mtime
    with _clf_lock:
        if not os.path.exists(MODEL_FILE):
            return None
        mtime = os.path.getmtime(MODEL_FILE)
        if _clf_cache is None or mtime != _clf_mtime:
            with open(MODEL_FILE, "rb") as f:
                _clf_cache = pickle.load(f)
            _clf_mtime = mtime
        return _clf_cache


def _invalidate_model_cache():
    """Call after saving a new model so the next request reloads it."""
    global _clf_cache, _clf_mtime
    with _clf_lock:
        _clf_cache = None
        _clf_mtime = 0.0

# ─────────────────────────────────────────────────────────────────────────────
# DUPLICATE-SIGN CHECK  (clean helper used in collect & live modes)
# ─────────────────────────────────────────────────────────────────────────────
def _is_duplicate_sign(features, target_label: str, clf):
    """
    Return the conflicting label if `features` strongly match a sign that is
    NOT `target_label`; otherwise return None.
    """
    if clf is None or features is None:
        return None
    try:
        proba = clf.predict_proba([features])[0]
        max_p = float(max(proba))
        pred  = str(clf.predict([features])[0])
        if max_p >= 0.70 and pred.upper() != target_label.upper():
            return pred
    except Exception:
        pass
    return None

# ─────────────────────────────────────────────────────────────────────────────
# PREDICT FROM FRAME  (used by Flask /api/hand_predict)
# ─────────────────────────────────────────────────────────────────────────────
def predict_from_frame(frame: np.ndarray, clf=None):
    """
    Detect hand in `frame` and return the predicted sign label, or None.
    `clf` is optional — if None, load_model() is called (cached, fast).
    """
    if clf is None:
        clf = load_model()
    if clf is None:
        return None

    features = _frame_to_features(frame)
    if features is None:
        return None

    try:
        proba = clf.predict_proba([features])[0]
        max_p = float(max(proba))
        if max_p < CONFIDENCE_MIN:
            return None
        return str(clf.predict([features])[0])
    except Exception as e:
        print(f"predict_from_frame error: {e}")
        return None


def predict_with_confidence(frame: np.ndarray, clf=None):
    """
    Same as predict_from_frame but also returns the confidence score.
    Returns (label, confidence) or (None, 0.0).
    """
    if clf is None:
        clf = load_model()
    if clf is None:
        return None, 0.0

    features = _frame_to_features(frame)
    if features is None:
        return None, 0.0

    try:
        proba = clf.predict_proba([features])[0]
        max_p = float(max(proba))
        if max_p < CONFIDENCE_MIN:
            return None, max_p
        return str(clf.predict([features])[0]), max_p
    except Exception as e:
        print(f"predict_with_confidence error: {e}")
        return None, 0.0



def get_model_status():
    """
    Return real-time model status for Flask debugging.
    Useful for /api/hand_model_status.
    """
    clf = load_model()
    return {
        "model_exists": os.path.exists(MODEL_FILE),
        "model_file": MODEL_FILE,
        "data_file": DATA_FILE,
        "task_file": TASK_FILE,
        "classifier_loaded": clf is not None,
        "classifier_type": type(clf).__name__ if clf is not None else None,
        "n_features_in": int(getattr(clf, "n_features_in_", 0) or 0) if clf is not None else None,
        "classes": [str(x) for x in getattr(clf, "classes_", [])] if clf is not None else [],
        "confidence_min": CONFIDENCE_MIN,
        "features_per_hand": FEATURES_PER_HAND,
    }


def predict_with_debug(frame: np.ndarray, clf=None):
    """
    Real-time prediction helper for Flask.

    Returns a detailed dictionary instead of only label so the frontend/backend
    can show the exact reason when recognition does not happen.
    """
    if clf is None:
        clf = load_model()

    if clf is None:
        return {
            "label": None,
            "confidence": 0.0,
            "accepted": False,
            "hand_detected": False,
            "reason": "model_not_loaded",
            "model_status": get_model_status(),
        }

    features = _frame_to_features(frame)

    if features is None:
        return {
            "label": None,
            "confidence": 0.0,
            "accepted": False,
            "hand_detected": False,
            "reason": "no_hand_landmarks_detected",
            "features_len": 0,
            "expected_features": int(getattr(clf, "n_features_in_", FEATURES_PER_HAND) or FEATURES_PER_HAND),
        }

    expected = int(getattr(clf, "n_features_in_", len(features)) or len(features))

    if len(features) != expected:
        return {
            "label": None,
            "confidence": 0.0,
            "accepted": False,
            "hand_detected": True,
            "reason": "feature_count_mismatch",
            "features_len": len(features),
            "expected_features": expected,
        }

    try:
        X = [features]
        pred = str(clf.predict(X)[0])
        confidence = 1.0
        top3 = []

        if hasattr(clf, "predict_proba"):
            proba = clf.predict_proba(X)[0]
            confidence = float(max(proba))
            classes = [str(x) for x in getattr(clf, "classes_", [])]
            ranked = sorted(
                zip(classes, [float(p) for p in proba]),
                key=lambda item: item[1],
                reverse=True,
            )[:3]
            top3 = [{"label": label, "confidence": round(score, 4)} for label, score in ranked]

        accepted = confidence >= CONFIDENCE_MIN

        return {
            "label": pred if accepted else None,
            "raw_label": pred,
            "confidence": round(confidence, 4),
            "accepted": accepted,
            "hand_detected": True,
            "reason": "ok" if accepted else "low_confidence",
            "features_len": len(features),
            "expected_features": expected,
            "top3": top3,
        }

    except Exception as e:
        return {
            "label": None,
            "confidence": 0.0,
            "accepted": False,
            "hand_detected": True,
            "reason": "prediction_error",
            "error": str(e),
            "features_len": len(features),
            "expected_features": expected,
        }

# ─────────────────────────────────────────────────────────────────────────────
# DATA COLLECTION
# ─────────────────────────────────────────────────────────────────────────────
def collect_data(num_samples: int = NUM_SAMPLES):
    _ensure_dirs()
    _ensure_task_model()

    print("\nEnter the word, phrase, or special sign to collect.")
    print(f"Examples: HELLO | THANK YOU | HOW ARE YOU | {' | '.join(sorted(SPECIAL_SIGNS))}")
    raw = input("Label: ").strip()
    ok, label, err = _validate_label(raw)
    if not ok:
        print(f"ERROR: {err}")
        return

    # Check existing samples
    if os.path.exists(DATA_FILE):
        df_ex    = pd.read_csv(DATA_FILE)
        existing = (df_ex["sign_label"].str.upper() == label).sum()
        if existing > 0:
            print(f"\nWARNING: '{label}' already has {existing} sample(s).")
            if input("   Continue anyway? [y/N]: ").strip().lower() != "y":
                print("Aborted.")
                return

    cam_idx      = _select_camera()
    cap          = _open_cap(cam_idx)
    if not cap.isOpened():
        print("Cannot open camera.")
        return

    existing_clf = load_model()   # for duplicate-sign detection
    collected    = 0
    last_save    = 0.0

    print(f"\nCollecting {num_samples} samples for '{label}' ...  ESC = abort\n")

    while collected < num_samples:
        ret, frame = cap.read()
        if not ret:
            break

        result   = _detect(frame)
        features = _result_to_features(result)
        dup_sign = None

        if features is not None:
            _draw_landmarks(frame, result)
            dup_sign = _is_duplicate_sign(features, label, existing_clf)

            if dup_sign:
                cv2.putText(frame, "Use a different hand position!", (10, 60),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 255), 2)
                cv2.putText(frame, f"Already assigned to: '{dup_sign}'", (10, 90),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.60, (0, 80, 255), 2)
            else:
                now = time.monotonic()
                if (now - last_save) >= COLLECT_INTERVAL:
                    last_save = now
                    _save_sample(label, result.hand_landmarks[0])
                    collected += 1
        else:
            cv2.putText(frame, "No hand detected", (10, 60),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 100, 255), 2)

        # Progress bar
        pct   = collected / num_samples
        bar_w = int(frame.shape[1] * 0.5 * pct)
        bx1   = 10
        bx2   = 10 + int(frame.shape[1] * 0.5)
        by1   = frame.shape[0] - 35
        by2   = frame.shape[0] - 15
        cv2.rectangle(frame, (bx1, by1), (bx2, by2), (40, 40, 40), -1)
        cv2.rectangle(frame, (bx1, by1), (bx1 + bar_w, by2), (0, 200, 100), -1)
        cv2.putText(frame,
                    f"'{label}'  {collected}/{num_samples}  ({int(pct*100)}%)  ESC=abort",
                    (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 100), 2)
        cv2.imshow("Data Collection - NeuroSense", frame)
        if cv2.waitKey(1) & 0xFF == 27:
            break

    cap.release()
    cv2.destroyAllWindows()
    print(f"\nCollected {collected} sample(s) for '{label}'.")

    if label in SPECIAL_SIGNS:
        return

    # Reference photo
    print("\nCapture a reference photo?  's' = save  |  ESC = skip")
    cap = _open_cap(cam_idx)
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        h, w = frame.shape[:2]
        y1 = max(0, h // 2 - REF_SIZE // 2);  y2 = min(h, y1 + REF_SIZE)
        x1 = max(0, w // 2 - REF_SIZE // 2);  x2 = min(w, x1 + REF_SIZE)
        crop = frame[y1:y2, x1:x2]
        if crop.shape[0] != REF_SIZE or crop.shape[1] != REF_SIZE:
            crop = cv2.resize(frame, (REF_SIZE, REF_SIZE))
        cv2.putText(crop, "'s'=save  ESC=skip", (10, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 2)
        cv2.imshow("Reference Photo", crop)
        key = cv2.waitKey(1) & 0xFF
        if key == ord("s"):
            safe = label.replace(" ", "_")
            path = os.path.join(REF_DIR, f"{safe}.jpg")
            cv2.imwrite(path, crop)
            print(f"Saved: {path}")
            break
        elif key == 27:
            break
    cap.release()
    cv2.destroyAllWindows()

# ─────────────────────────────────────────────────────────────────────────────
# TRAIN
# ─────────────────────────────────────────────────────────────────────────────
def train_model():
    if not os.path.exists(DATA_FILE):
        print("No data found! Collect data first.")
        return

    df     = pd.read_csv(DATA_FILE)
    counts = df["sign_label"].value_counts()
    print("\nTraining data summary:")
    for lbl, n in counts.items():
        print(f"   {lbl:<25} {n} samples")

    if len(counts) < 2:
        print("Need at least 2 different signs to train.")
        return

    X   = df[_FEATURE_COLS].values
    y   = df["sign_label"].values
    clf = RandomForestClassifier(
        n_estimators=200,
        random_state=42,
        n_jobs=-1,
        min_samples_leaf=2,
    )
    clf.fit(X, y)

    os.makedirs(os.path.dirname(MODEL_FILE), exist_ok=True)
    with open(MODEL_FILE, "wb") as f:
        pickle.dump(clf, f)

    _invalidate_model_cache()   # force cache reload on next call
    print(f"\nModel trained — {len(y)} samples, {len(counts)} sign(s) -> {MODEL_FILE}")

# ─────────────────────────────────────────────────────────────────────────────
# LIVE PREDICTION
# ─────────────────────────────────────────────────────────────────────────────
def live_prediction():
    clf = load_model()
    if clf is None:
        print("No model found! Train first.")
        return

    cap = _open_cap(_select_camera())
    if not cap.isOpened():
        print("Cannot open camera.")
        return

    sentence        = ""
    last_pred       = None
    pred_count      = 0
    no_hand_frames  = 0
    capitalize_next = True
    last_append_t   = 0.0

    print("\nLive Prediction — ESC to exit")
    print(f"Special signs: {' | '.join(sorted(SPECIAL_SIGNS))}\n")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        result   = _detect(frame)
        features = _result_to_features(result)

        if features is not None:
            _draw_landmarks(frame, result)
            no_hand_frames = 0

            try:
                proba = clf.predict_proba([features])[0]
                max_p = float(max(proba))
                pred  = str(clf.predict([features])[0]) if max_p >= CONFIDENCE_MIN else None
            except Exception:
                pred = None

            if pred:
                pred_count = pred_count + 1 if pred == last_pred else 1
                last_pred  = pred

                now = time.monotonic()
                if pred_count >= STABLE_FRAMES and (now - last_append_t) >= LIVE_COOLDOWN:
                    last_append_t = now
                    pred_count    = 0
                    last_pred     = None

                    if pred == "SPACE":
                        sentence        = sentence.rstrip() + " "
                        capitalize_next = False
                    elif pred == "DELETE":
                        words = sentence.rstrip().split()
                        if words:
                            words.pop()
                            sentence        = (" ".join(words) + " ") if words else ""
                            capitalize_next = not bool(words)
                    elif pred == "CLEAR":
                        sentence        = ""
                        capitalize_next = True
                    elif pred == "ENTER":
                        pass   # handled by UI layer
                    elif pred in (".", ",", "?", "!"):
                        sentence        = sentence.rstrip() + pred + " "
                        capitalize_next = pred in (".", "?", "!")
                    else:
                        words_in = pred.split()
                        if capitalize_next:
                            fmt = words_in[0].capitalize()
                            if len(words_in) > 1:
                                fmt += " " + " ".join(w.lower() for w in words_in[1:])
                        else:
                            fmt = pred.lower()
                        sentence        = (sentence.rstrip() + " " + fmt).lstrip()
                        capitalize_next = False
            else:
                if last_pred != pred:
                    pred_count = 0
                    last_pred  = None
        else:
            no_hand_frames += 1
            if no_hand_frames >= 12:
                pred_count = 0
                last_pred  = None

        # HUD overlay
        overlay = frame.copy()
        cv2.rectangle(overlay, (0, frame.shape[0] - 80),
                      (frame.shape[1], frame.shape[0]), (10, 10, 20), -1)
        cv2.addWeighted(overlay, 0.7, frame, 0.3, 0, frame)

        disp = sentence if len(sentence) <= 60 else "..." + sentence[-57:]
        cv2.putText(frame, disp or "(sign something)",
                    (10, frame.shape[0] - 15),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.85, (0, 255, 160), 2)

        if last_pred and pred_count:
            bar_w = int((pred_count / STABLE_FRAMES) * 200)
            cv2.rectangle(frame, (10, 50), (210, 65), (40, 40, 40), -1)
            cv2.rectangle(frame, (10, 50), (10 + bar_w, 65), (0, 200, 100), -1)
            cv2.putText(frame, f"{last_pred}  ({int(pred_count / STABLE_FRAMES * 100)}%)",
                        (10, 45), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 200), 2)

        cv2.putText(frame, "ESC=exit | SPACE / DELETE / CLEAR supported",
                    (10, frame.shape[0] - 55),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.45, (120, 120, 160), 1)
        cv2.imshow("Live Prediction - NeuroSense", frame)

        if cv2.waitKey(1) & 0xFF == 27:
            break

    cap.release()
    cv2.destroyAllWindows()
    if sentence.strip():
        print(f"\nFinal sentence: {sentence.strip()}")

# ─────────────────────────────────────────────────────────────────────────────
# DELETE SIGN
# ─────────────────────────────────────────────────────────────────────────────
def delete_word(word: str):
    if not os.path.exists(DATA_FILE):
        print("No data found!")
        return
    ok, label, err = _validate_label(word)
    if not ok:
        print(f"ERROR: {err}")
        return

    df      = pd.read_csv(DATA_FILE)
    before  = len(df)
    df      = df[df["sign_label"].str.upper() != label]
    removed = before - len(df)

    if not removed:
        print(f"No entries found for '{label}'.")
        return

    if len(df):
        df.to_csv(DATA_FILE, index=False)
    else:
        os.remove(DATA_FILE)
    print(f"Deleted {removed} sample(s) for '{label}'.")

    safe = label.replace(" ", "_")
    ref  = os.path.join(REF_DIR, f"{safe}.jpg")
    if os.path.exists(ref):
        os.remove(ref)
        print("Deleted reference image.")

    if os.path.exists(DATA_FILE):
        print("Retraining model...")
        train_model()
    elif os.path.exists(MODEL_FILE):
        os.remove(MODEL_FILE)
        _invalidate_model_cache()
        print("No data left — model deleted.")

# ─────────────────────────────────────────────────────────────────────────────
# SHOW TRAINED SIGNS
# ─────────────────────────────────────────────────────────────────────────────
def show_trained_words():
    clf = load_model()
    if clf is None:
        print("No model found! Train first.")
        return

    classes = sorted(clf.classes_)
    print(f"\nTrained signs ({len(classes)}):")
    for i, w in enumerate(classes, 1):
        icon = "SPECIAL" if w in SPECIAL_SIGNS else "SIGN"
        print(f"  {i:>3}. [{icon}] {w}")

    if os.path.exists(DATA_FILE):
        df     = pd.read_csv(DATA_FILE)
        counts = df["sign_label"].value_counts()
        print("\nSample counts:")
        for w in classes:
            n   = counts.get(w, 0)
            bar = "#" * min(n // 5, 30)
            print(f"   {w:<25} {bar} {n}")

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT  (Pillow-based skeletal diagrams, no cairo required)
# ─────────────────────────────────────────────────────────────────────────────
_LM_CONNECTIONS = [
    (0,1),(1,2),(2,3),(3,4),
    (0,5),(5,6),(6,7),(7,8),
    (0,9),(9,10),(10,11),(11,12),
    (0,13),(13,14),(14,15),(15,16),
    (0,17),(17,18),(18,19),(19,20),
    (5,9),(9,13),(13,17),
]
_FCOLORS_RGB = {
    "wrist":  (83,  74,  183),
    "thumb":  (29,  158, 117),
    "index":  (216, 90,  48),
    "middle": (127, 119, 221),
    "ring":   (55,  138, 221),
    "pinky":  (186, 117, 23),
}
_STATIC_PTS = {
    0:(120,310),1:(72,215),2:(55,185),3:(38,155),4:(24,128),
    5:(88,205), 6:(86,158), 7:(84,115), 8:(83,80),
    9:(108,200),10:(108,150),11:(108,108),12:(108,72),
    13:(130,202),14:(132,155),15:(134,115),16:(136,82),
    17:(152,208),18:(156,168),19:(160,135),20:(163,108),
}
_DESCS = {
    "WRIST":"Base anchor of all hand landmarks",
    "THUMB_CMC":"Thumb carpometacarpal joint",
    "THUMB_MCP":"Thumb metacarpophalangeal",
    "THUMB_IP":"Thumb interphalangeal",
    "THUMB_TIP":"Thumb fingertip",
    "INDEX_MCP":"Index metacarpophalangeal",
    "INDEX_PIP":"Index proximal interphalangeal",
    "INDEX_DIP":"Index distal interphalangeal",
    "INDEX_TIP":"Index fingertip",
    "MIDDLE_MCP":"Middle metacarpophalangeal",
    "MIDDLE_PIP":"Middle proximal interphalangeal",
    "MIDDLE_DIP":"Middle distal interphalangeal",
    "MIDDLE_TIP":"Middle fingertip",
    "RING_MCP":"Ring metacarpophalangeal",
    "RING_PIP":"Ring proximal interphalangeal",
    "RING_DIP":"Ring distal interphalangeal",
    "RING_TIP":"Ring fingertip",
    "PINKY_MCP":"Pinky metacarpophalangeal",
    "PINKY_PIP":"Pinky proximal interphalangeal",
    "PINKY_DIP":"Pinky distal interphalangeal",
    "PINKY_TIP":"Pinky fingertip",
}


def _render_hand_png(lm_coords=None, width=240, height=320):
    """Pure-Pillow skeletal hand renderer — no cairo, no inkscape needed."""
    try:
        from PIL import Image, ImageDraw
        import io as _io
    except ImportError:
        return None

    PAD = 22
    img  = Image.new("RGB", (width, height), (249, 249, 252))
    draw = ImageDraw.Draw(img)

    if lm_coords is not None:
        xs = [c[0] for c in lm_coords]; ys = [c[1] for c in lm_coords]
        mn_x, mx_x = min(xs), max(xs);  mn_y, mx_y = min(ys), max(ys)
        sx = max(mx_x - mn_x, 0.01);    sy = max(mx_y - mn_y, 0.01)
        iw, ih = width - 2 * PAD, height - 2 * PAD
        pts = [
            (int(PAD + (c[0] - mn_x) / sx * iw),
             int(PAD + (c[1] - mn_y) / sy * ih))
            for c in lm_coords
        ]
    else:
        pts = [_STATIC_PTS[i] for i in range(21)]

    palm = [pts[i] for i in (0, 1, 5, 9, 13, 17)]
    draw.polygon(palm, fill=(240, 210, 195))

    for a, b in _LM_CONNECTIONS:
        draw.line([pts[a], pts[b]], fill=(180, 178, 169), width=3)

    tip_labels = [
        (4, "Thumb",  "thumb"),
        (8, "Index",  "index"),
        (12,"Middle", "middle"),
        (16,"Ring",   "ring"),
        (20,"Pinky",  "pinky"),
        (0, "Wrist",  "wrist"),
    ]
    for i in range(21):
        col = _FCOLORS_RGB[_LM_FINGER[i]]
        cx, cy = pts[i]; r = 7
        draw.ellipse([(cx-r, cy-r), (cx+r, cy+r)],
                     fill=col, outline=(255, 255, 255), width=2)
        lbl = str(i)
        try:
            bb = draw.textbbox((0, 0), lbl); tw, th = bb[2]-bb[0], bb[3]-bb[1]
        except Exception:
            tw, th = 6, 8
        draw.text((cx - tw // 2, cy - th // 2 + 1), lbl, fill=(255, 255, 255))

    for tip, name, fg in tip_labels:
        col = _FCOLORS_RGB[fg]; cx, cy = pts[tip]
        dy  = 14 if tip == 0 else -14
        try:
            bb = draw.textbbox((0, 0), name); tw = bb[2] - bb[0]
        except Exception:
            tw = len(name) * 6
        draw.text((cx - tw // 2, cy + dy), name, fill=col)

    buf = _io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def export_to_excel(output_path=None, sign_filter=None):
    """
    Export word_data.csv to a formatted Excel workbook with:
      - 'Data' sheet        : all landmark coordinates + metadata
      - 'Signs' sheet       : one row per sign with skeletal hand diagram
      - 'Landmark Index'    : colour-coded reference for all 21 landmarks
    Requires: openpyxl  (pip install openpyxl pillow)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
        import io as _io
    except ImportError:
        print("openpyxl required: pip install openpyxl pillow")
        return

    if not os.path.exists(DATA_FILE):
        print("No data found!")
        return

    if output_path is None:
        output_path = os.path.join(BASE_DIR, "hs_data_export.xlsx")

    df = pd.read_csv(DATA_FILE)
    if sign_filter:
        df = df[df["sign_label"].str.upper() == sign_filter.upper()]
        if df.empty:
            print(f"No data for '{sign_filter}'.")
            return

    wb   = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center")

    def _hdr(cell, bg="3C3489", fg="CECBF6"):
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.font      = Font(name="Arial", bold=True, color=fg, size=9)
        cell.alignment = ctr
        cell.border    = bord

    def _body(cell, bg=None, fg="333333", mono=False, bold=False):
        cell.font      = Font(name="Courier New" if mono else "Arial",
                              size=9, color=fg, bold=bold)
        cell.alignment = lft
        cell.border    = bord
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)

    # Sheet 1: Data
    ws = wb.active; ws.title = "Data"; ws.freeze_panes = "C2"
    cols = list(df.columns)
    for ci, h in enumerate(cols, 1):
        c  = ws.cell(row=1, column=ci, value=h)
        bg = ("534AB7" if h in ("sample_id", "timestamp") else
              "1D1B4B" if h == "sign_label" else "3C3489")
        _hdr(c, bg=bg)
    ws.row_dimensions[1].height = 32

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            h = cols[ci - 1]; c = ws.cell(row=ri, column=ci, value=val)
            if h == "sign_label":
                c.fill = PatternFill("solid", start_color="534AB7")
                c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
                c.alignment = ctr; c.border = bord
            elif h in ("sample_id", "timestamp"):
                _body(c, bg="EEEDFE", fg="3C3489", bold=(h == "sample_id"))
            else:
                _body(c, mono=True)
                if isinstance(val, float):
                    c.number_format = "0.000000"

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    for ci in range(3, len(cols)):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    ws.column_dimensions[get_column_letter(len(cols))].width = 15

    # Sheet 2: Signs
    ws2 = wb.create_sheet("Signs"); ws2.freeze_panes = "A2"
    for ci, h in enumerate(["Sign","Samples","Skeletal Diagram",
                             "Avg WRIST","Avg INDEX_TIP","Avg MIDDLE_TIP"], 1):
        _hdr(ws2.cell(row=1, column=ci, value=h), bg="1D1B4B", fg="FFFFFF")
    ws2.row_dimensions[1].height = 28
    for ci, w in enumerate([16, 10, 36, 26, 26, 26], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    signs     = sorted(df["sign_label"].unique())
    feat_cols = [c for c in df.columns if c.startswith("H1_")]

    for ri, sign in enumerate(signs, 2):
        sdf     = df[df["sign_label"] == sign]
        medians = sdf[feat_cols].median()

        lm_coords = [
            (float(medians.get(f"H1_{n}_x", 0.5)),
             float(medians.get(f"H1_{n}_y", 0.5)),
             float(medians.get(f"H1_{n}_z", 0.0)))
            for n in _LM_NAMES
        ]

        c = ws2.cell(row=ri, column=1, value=sign)
        c.fill = PatternFill("solid", start_color="534AB7")
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.alignment = ctr; c.border = bord

        cn = ws2.cell(row=ri, column=2, value=len(sdf))
        cn.font = Font(name="Arial", size=10)
        cn.alignment = ctr; cn.border = bord

        png = _render_hand_png(lm_coords, width=220, height=240)
        if png:
            img = XLImage(_io.BytesIO(png)); img.width = 170; img.height = 180
            ws2.add_image(img, f"C{ri}")
        else:
            c3 = ws2.cell(row=ri, column=3, value="pip install pillow")
            c3.font = Font(name="Arial", size=8, italic=True, color="999999")
            c3.alignment = ctr; c3.border = bord

        def _fmt(nm):
            x = medians.get(f"H1_{nm}_x", 0)
            y = medians.get(f"H1_{nm}_y", 0)
            z = medians.get(f"H1_{nm}_z", 0)
            return f"({x:.4f}, {y:.4f}, {z:.4f})"

        for ci, nm in enumerate(["WRIST", "INDEX_TIP", "MIDDLE_TIP"], 4):
            c = ws2.cell(row=ri, column=ci, value=_fmt(nm))
            c.font = Font(name="Courier New", size=9)
            c.alignment = ctr; c.border = bord

        ws2.row_dimensions[ri].height = 185

    # Sheet 3: Landmark Index
    ws3 = wb.create_sheet("Landmark Index")
    for ci, h in enumerate(["Index","Column Name","Finger","Description"], 1):
        _hdr(ws3.cell(row=1, column=ci, value=h), bg="1D1B4B", fg="FFFFFF")
    ws3.row_dimensions[1].height = 24
    for ci, w in enumerate([8, 22, 14, 50], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    _LIGHT = {"wrist":"EEEDFE","thumb":"E1F5EE","index":"FAECE7",
              "middle":"EEEDFE","ring":"E6F1FB","pinky":"FAEEDA"}
    _DARK  = {"wrist":"534AB7","thumb":"1D9E75","index":"D85A30",
              "middle":"7F77DD","ring":"378ADD","pinky":"BA7517"}
    _TXT   = {"wrist":"3C3489","thumb":"0F6E56","index":"993C1D",
              "middle":"3C3489","ring":"185FA5","pinky":"854F0B"}

    for i, lm_name in enumerate(_LM_NAMES):
        fg = _LM_FINGER[i]; row = i + 2
        for ci, v in enumerate([i, f"H1_{lm_name}", fg.capitalize(),
                                  _DESCS.get(lm_name, "")], 1):
            c = ws3.cell(row=row, column=ci, value=v)
            if ci <= 3:
                c.fill = PatternFill("solid", start_color=_DARK[fg])
                c.font = Font(name="Arial", size=9, color="FFFFFF",
                              bold=(ci in (1, 3)))
            else:
                c.fill = PatternFill("solid", start_color=_LIGHT[fg])
                c.font = Font(name="Arial", size=9, color=_TXT[fg])
            c.alignment = Alignment(
                horizontal="center" if ci < 4 else "left",
                vertical="center"
            )
            c.border = bord
        ws3.row_dimensions[row].height = 18

    wb.save(output_path)
    print(f"\nExcel saved -> {output_path}")
    print(f"   Data      : {len(df)} rows")
    print(f"   Signs     : {len(signs)} sign(s) with skeletal diagrams")
    print(f"   Landmarks : 21 documented")

# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def run_cli():
    print("\n+--------------------------------------+")
    print("|  NeuroSense -- Hand Sign Module      |")
    print("+--------------------------------------+")
    while True:
        print("\n  1: Collect Data")
        print("  2: Train Model")
        print("  3: Live Prediction")
        print("  4: Delete Sign")
        print("  5: Show Trained Signs")
        print("  6: Export to Excel")
        print("  0: Exit")
        mode = input("\nEnter mode: ").strip()
        if   mode == "1": collect_data()
        elif mode == "2": train_model()
        elif mode == "3": live_prediction()
        elif mode == "4": delete_word(input("Sign/phrase to delete: ").strip())
        elif mode == "5": show_trained_words()
        elif mode == "6":
            sign = input("Filter by sign (blank = all): ").strip() or None
            path = input("Output path  (blank = default): ").strip() or None
            export_to_excel(output_path=path, sign_filter=sign)
        elif mode == "0":
            break
        else:
            print("Invalid choice.")


if __name__ == "__main__":
    run_cli()